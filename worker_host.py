import argparse
import json
import tempfile
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

from onpe_core import OnpeBrowserClient, cell_to_str, normalize_header


PREVIEW_LIMIT = 20


def format_elapsed(delta):
    total_seconds = int(delta.total_seconds())
    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f'{hours}h {minutes}m {seconds}s'
    if minutes:
        return f'{minutes}m {seconds}s'
    return f'{seconds}s'


def load_records(path):
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    raw_headers = [cell.value if cell.value is not None else '' for cell in ws[1]]

    headers = {}
    for index, header in enumerate(raw_headers, start=1):
        normalized = normalize_header(header)
        if normalized:
            headers[normalized] = index

    if 'dni' not in headers:
        wb.close()
        raise ValueError('El Excel debe tener una columna llamada dni.')

    records = []
    for row_number in range(2, ws.max_row + 1):
        dni = ws.cell(row=row_number, column=headers['dni']).value
        dni = '' if dni is None else str(dni).strip()
        if not dni:
            continue

        record = {'row_number': row_number, 'dni': dni, 'estado': 'pendiente', 'detalle_error': ''}
        for header_norm, col_idx in headers.items():
            record[header_norm] = cell_to_str(ws.cell(row=row_number, column=col_idx).value)
        records.append(record)

    wb.close()
    return headers, records


def write_record(ws, headers, record):
    for header_norm, col_idx in headers.items():
        if header_norm in record:
            ws.cell(row=record['row_number'], column=col_idx, value=record.get(header_norm, ''))


def process_excel(path):
    headers, records = load_records(path)
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    total = len(records)
    errors = 0

    with OnpeBrowserClient() as client:
        for index, record in enumerate(records, start=1):
            print(f'Consultando DNI {index}/{total}: {record["dni"]}', flush=True)
            try:
                result = client.consultar_dni(record['dni'])
                for key, value in result.values.items():
                    if key in headers:
                        record[key] = value
                record['estado'] = 'consultado'
                record['detalle_error'] = ''
            except Exception as exc:
                errors += 1
                record['estado'] = 'error'
                record['detalle_error'] = str(exc)
                if 'miembro_de_mesa' in headers:
                    record['miembro_de_mesa'] = 'error'

            write_record(ws, headers, record)

    wb.save(path)
    wb.close()
    return records, total, errors


def build_preview(records):
    preview = []
    for record in records[:PREVIEW_LIMIT]:
        preview.append(
            {
                'dni': record.get('dni', ''),
                'miembro_de_mesa': record.get('miembro_de_mesa', ''),
                'ubicacion': record.get('ubicacion', ''),
                'direccion': record.get('direccion', ''),
                'estado': record.get('estado', ''),
                'detalle_error': record.get('detalle_error', ''),
            }
        )
    return preview


def fetch_next_job(session, server_url):
    response = session.post(f'{server_url}/api/jobs/next', timeout=30)
    response.raise_for_status()
    return response.json().get('job')


def report_error(session, server_url, job_id, message):
    session.post(
        f'{server_url}/api/jobs/{job_id}/error',
        json={'message': message},
        timeout=30,
    )


def complete_job(session, server_url, job_id, result_path, total, errors, elapsed, records):
    with open(result_path, 'rb') as handle:
        response = session.post(
            f'{server_url}/api/jobs/{job_id}/complete',
            data={
                'total': str(total),
                'errors': str(errors),
                'elapsed': elapsed,
                'message': 'Proceso completado por el worker local.',
                'records_json': json.dumps(build_preview(records), ensure_ascii=False),
            },
            files={'result_file': (Path(result_path).name, handle, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            timeout=120,
        )
    response.raise_for_status()


def run_worker(server_url, interval, once):
    session = requests.Session()
    print(f'Worker conectado a {server_url}', flush=True)

    while True:
        job = fetch_next_job(session, server_url)
        if not job:
            print('No hay trabajos pendientes.', flush=True)
            if once:
                break
            time.sleep(interval)
            continue

        job_id = job['id']
        print(f'Trabajo recibido: {job_id} | {job["original_name"]}', flush=True)

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_dir_path = Path(temp_dir)
            local_file = temp_dir_path / job['original_name']
            try:
                download = session.get(job['download_url'], timeout=120)
                download.raise_for_status()
                local_file.write_bytes(download.content)

                started = datetime.now()
                records, total, errors = process_excel(local_file)
                elapsed = format_elapsed(datetime.now() - started)
                complete_job(session, server_url, job_id, local_file, total, errors, elapsed, records)
                print(f'Trabajo completado: {job_id} en {elapsed}', flush=True)
            except Exception as exc:
                report_error(session, server_url, job_id, str(exc))
                print(f'Error en trabajo {job_id}: {exc}', flush=True)
                if once:
                    raise

        if once:
            break


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Worker local para procesar ONPE fuera de Docker.')
    parser.add_argument('--server', default='http://localhost:8000', help='URL base de la web dockerizada.')
    parser.add_argument('--interval', type=int, default=5, help='Segundos entre consultas cuando no hay trabajos pendientes.')
    parser.add_argument('--once', action='store_true', help='Procesa un solo trabajo y termina.')
    args = parser.parse_args()

    run_worker(args.server.rstrip('/'), args.interval, args.once)
