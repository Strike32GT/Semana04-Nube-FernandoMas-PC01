import os
import threading
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import load_workbook

from onpe_core import DEFAULT_EXCEL_NAME, OnpeBrowserClient, cell_to_str, normalize_header


ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class ConsultaElectoralApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Consulta ONPE - Excel")
        self.geometry("1120x780")
        self.minsize(980, 700)

        self.archivo_excel = os.path.join(os.getcwd(), DEFAULT_EXCEL_NAME)
        self.registros = []
        self.headers = {}
        self.display_headers = {}
        self.procesando = False

        self._build_ui()
        self._cargar_excel_inicial()

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        main = ctk.CTkFrame(self, corner_radius=16)
        main.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(5, weight=1)

        ctk.CTkLabel(
            main,
            text="Consulta masiva de miembros de mesa",
            font=ctk.CTkFont(size=30, weight="bold"),
        ).grid(row=0, column=0, padx=20, pady=(20, 8), sticky="w")

        ctk.CTkLabel(
            main,
            text=(
                "Carga un Excel con la columna dni. El programa visitara ONPE en segundo plano, "
                "llenara las columnas reconocidas de tu Excel y luego mostrara la vista previa en la app."
            ),
            wraplength=980,
            justify="left",
            font=ctk.CTkFont(size=15),
        ).grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")

        top = ctk.CTkFrame(main)
        top.grid(row=2, column=0, padx=20, pady=(0, 12), sticky="ew")
        top.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(top, text="Cargar Excel", width=150, height=40, command=self.seleccionar_excel).grid(
            row=0, column=0, padx=12, pady=12
        )

        self.archivo_entry = ctk.CTkEntry(top, height=40)
        self.archivo_entry.grid(row=0, column=1, padx=(0, 12), pady=12, sticky="ew")
        self.archivo_entry.insert(0, self.archivo_excel)

        self.procesar_button = ctk.CTkButton(top, text="Consultar ONPE", width=170, height=40, command=self.procesar_excel)
        self.procesar_button.grid(row=0, column=2, padx=(0, 12), pady=12)

        self.archivo_status = ctk.CTkLabel(main, text="Archivo: sin cargar", anchor="w", justify="left", font=ctk.CTkFont(size=15, weight="bold"))
        self.archivo_status.grid(row=3, column=0, padx=20, pady=(0, 8), sticky="ew")

        self.progreso_label = ctk.CTkLabel(main, text="Progreso: esperando archivo", anchor="w", justify="left", font=ctk.CTkFont(size=14))
        self.progreso_label.grid(row=4, column=0, padx=20, pady=(0, 12), sticky="ew")

        self.lista_frame = ctk.CTkScrollableFrame(main)
        self.lista_frame.grid(row=5, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.lista_frame.grid_columnconfigure(0, weight=1)

    def _cargar_excel_inicial(self):
        if os.path.exists(self.archivo_excel):
            try:
                self.cargar_excel(self.archivo_excel)
            except Exception as exc:
                self.archivo_status.configure(text=f"Archivo: no se pudo cargar el Excel inicial ({exc})", text_color="#B91C1C")

    def seleccionar_excel(self):
        path = filedialog.askopenfilename(title="Seleccionar archivo Excel", filetypes=[("Archivos de Excel", "*.xlsx")])
        if not path:
            return
        try:
            self.cargar_excel(path)
            messagebox.showinfo("Excel cargado", "El archivo se cargo correctamente. La consulta iniciara automaticamente.")
            self.procesar_excel()
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo cargar el archivo Excel.\n\n{exc}")

    def cargar_excel(self, path):
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        raw_headers = [cell.value if cell.value is not None else "" for cell in ws[1]]

        self.headers = {}
        self.display_headers = {}
        for index, header in enumerate(raw_headers, start=1):
            normalized = normalize_header(header)
            if normalized:
                self.headers[normalized] = index
                self.display_headers[normalized] = str(header).strip()

        if "dni" not in self.headers:
            wb.close()
            raise ValueError("El Excel debe tener una columna llamada dni.")

        registros = []
        for row_number in range(2, ws.max_row + 1):
            dni = ws.cell(row=row_number, column=self.headers["dni"]).value
            dni = "" if dni is None else str(dni).strip()
            if not dni:
                continue

            registro = {"row_number": row_number, "dni": dni, "estado": "pendiente", "detalle_error": ""}
            for header_norm, col_idx in self.headers.items():
                registro[header_norm] = cell_to_str(ws.cell(row=row_number, column=col_idx).value)
            registros.append(registro)

        wb.close()

        self.archivo_excel = path
        self.registros = registros
        self.archivo_entry.delete(0, "end")
        self.archivo_entry.insert(0, path)
        self.archivo_status.configure(text=f"Archivo cargado: {os.path.basename(path)} | Registros: {len(registros)}", text_color="#166534")
        self.progreso_label.configure(text="Progreso: listo para consultar ONPE", text_color="#334155")
        self._render_registros()

    def procesar_excel(self):
        if self.procesando:
            return
        if not self.registros:
            return

        self.procesando = True
        self.procesar_button.configure(state="disabled", text="Consultando...")
        self.progreso_label.configure(text="Progreso: consultando ONPE en segundo plano...", text_color="#1D4ED8")
        threading.Thread(target=self._procesar_excel_worker, daemon=True).start()

    def _procesar_excel_worker(self):
        errores = 0
        try:
            wb = load_workbook(self.archivo_excel)
            ws = wb[wb.sheetnames[0]]
            total = len(self.registros)

            with OnpeBrowserClient() as client:
                for index, registro in enumerate(self.registros, start=1):
                    try:
                        resultado = client.consultar_dni(registro["dni"])
                        self._apply_result_to_record(registro, resultado)
                        self._write_record_to_sheet(ws, registro)
                    except Exception as exc:
                        errores += 1
                        registro["estado"] = "error"
                        registro["detalle_error"] = str(exc)
                        if "miembro_de_mesa" in self.headers:
                            registro["miembro_de_mesa"] = "error"
                            ws.cell(row=registro["row_number"], column=self.headers["miembro_de_mesa"], value="error")

                    self.after(0, self._actualizar_progreso, index, total, registro["dni"], registro.get("estado", ""), registro.get("detalle_error", ""))

            ruta_guardada = self._guardar_workbook(wb)
            wb.close()
            self.after(0, self._finalizar_proceso_exitoso, total, errores, ruta_guardada)
        except Exception as exc:
            self.after(0, self._finalizar_proceso_con_error, str(exc))

    def _apply_result_to_record(self, registro, resultado):
        for key, value in resultado.values.items():
            if key in self.headers:
                registro[key] = value
        registro["estado"] = "consultado"
        registro["detalle_error"] = resultado.error

    def _write_record_to_sheet(self, ws, registro):
        for header_norm, col_idx in self.headers.items():
            if header_norm in registro:
                ws.cell(row=registro["row_number"], column=col_idx, value=registro.get(header_norm, ""))

    def _guardar_workbook(self, workbook):
        try:
            workbook.save(self.archivo_excel)
            return self.archivo_excel
        except PermissionError:
            raise RuntimeError(
                "No se pudo sobrescribir el Excel original porque esta abierto o bloqueado. "
                "Cierralo y vuelve a intentar."
            )

    def _actualizar_progreso(self, actual, total, dni, estado, detalle_error):
        texto = f"Progreso: {actual}/{total} | DNI {dni} -> {estado}"
        if detalle_error:
            texto += f" | {detalle_error}"
        self.progreso_label.configure(text=texto, text_color="#1D4ED8" if estado != "error" else "#B91C1C")
        self._render_registros()

    def _finalizar_proceso_exitoso(self, total, errores, ruta_guardada):
        self.procesando = False
        self.procesar_button.configure(state="normal", text="Consultar ONPE")
        self.progreso_label.configure(
            text=f"Progreso: proceso completado. Se actualizaron {total} registros, con {errores} errores, y se guardo el Excel.",
            text_color="#166534",
        )
        self._render_registros()

    def _finalizar_proceso_con_error(self, error_message):
        self.procesando = False
        self.procesar_button.configure(state="normal", text="Consultar ONPE")
        self.progreso_label.configure(text=f"Progreso: ocurrio un error durante la consulta. {error_message}", text_color="#B91C1C")

    def _render_registros(self):
        for widget in self.lista_frame.winfo_children():
            widget.destroy()

        if not self.registros:
            ctk.CTkLabel(self.lista_frame, text="No hay registros para mostrar.", font=ctk.CTkFont(size=18, weight="bold")).grid(
                row=0, column=0, padx=20, pady=30, sticky="w"
            )
            return

        for index, registro in enumerate(self.registros):
            self._crear_card_registro(index, registro)

    def _crear_card_registro(self, row_index, registro):
        card = ctk.CTkFrame(self.lista_frame, corner_radius=14)
        card.grid(row=row_index, column=0, padx=10, pady=10, sticky="ew")
        card.grid_columnconfigure(0, weight=1)
        card.grid_columnconfigure(1, weight=1)

        miembro_valor = registro.get("miembro_de_mesa", "")
        status_text = miembro_valor.upper() if miembro_valor else registro.get("estado", "pendiente").upper()
        if status_text == "SI":
            badge_color = "#15803D"
        elif status_text == "NO":
            badge_color = "#DC2626"
        elif status_text == "ERROR":
            badge_color = "#B91C1C"
        else:
            badge_color = "#F59E0B"

        header = ctk.CTkFrame(card, fg_color="transparent")
        header.grid(row=0, column=0, columnspan=2, padx=16, pady=(16, 10), sticky="ew")
        header.grid_columnconfigure(1, weight=1)

        badge = ctk.CTkFrame(header, fg_color=badge_color, corner_radius=12)
        badge.grid(row=0, column=0, padx=(0, 12), sticky="w")
        ctk.CTkLabel(badge, text=f"MIEMBRO DE MESA: {status_text}", text_color="white", font=ctk.CTkFont(size=15, weight="bold")).pack(
            padx=16, pady=12
        )

        ctk.CTkLabel(header, text=f"DNI: {registro['dni']}", justify="left", anchor="w", font=ctk.CTkFont(size=18, weight="bold")).grid(
            row=0, column=1, sticky="ew"
        )

        self._add_detail(card, 1, 0, "Ubicacion", registro.get("ubicacion", "") or "Pendiente de consultar ONPE")
        self._add_detail(card, 1, 1, "Direccion", registro.get("direccion", "") or "Pendiente de consultar ONPE")

        extras = self._extra_fields_for_preview(registro)
        if extras:
            extra_frame = ctk.CTkFrame(card)
            extra_frame.grid(row=2, column=0, columnspan=2, padx=12, pady=(0, 16), sticky="ew")
            extra_frame.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(extra_frame, text="Campos adicionales", anchor="w", font=ctk.CTkFont(size=14, weight="bold")).pack(
                fill="x", padx=14, pady=(12, 6)
            )
            ctk.CTkLabel(
                extra_frame,
                text=" | ".join(f"{label}: {value}" for label, value in extras),
                justify="left",
                wraplength=920,
                anchor="w",
                font=ctk.CTkFont(size=14),
            ).pack(fill="x", padx=14, pady=(0, 12))

        if registro.get("detalle_error"):
            error_frame = ctk.CTkFrame(card)
            error_frame.grid(row=3, column=0, columnspan=2, padx=12, pady=(0, 16), sticky="ew")
            ctk.CTkLabel(error_frame, text=f"Detalle: {registro['detalle_error']}", text_color="#B91C1C", anchor="w", justify="left", wraplength=920).pack(
                fill="x", padx=14, pady=12
            )

    def _extra_fields_for_preview(self, registro):
        hidden = {"row_number", "dni", "miembro_de_mesa", "ubicacion", "direccion", "estado", "detalle_error"}
        extras = []
        for header_norm in self.headers:
            if header_norm in hidden:
                continue
            value = registro.get(header_norm, "")
            if value:
                extras.append((self.display_headers.get(header_norm, header_norm), value))
        return extras

    def _add_detail(self, parent, row, col, title, value):
        frame = ctk.CTkFrame(parent)
        frame.grid(row=row, column=col, padx=12, pady=(0, 16), sticky="nsew")
        ctk.CTkLabel(frame, text=title, anchor="w", font=ctk.CTkFont(size=15, weight="bold")).pack(fill="x", padx=14, pady=(12, 4))
        ctk.CTkLabel(frame, text=value, anchor="w", justify="left", wraplength=430, font=ctk.CTkFont(size=16)).pack(fill="x", padx=14, pady=(0, 12))


if __name__ == "__main__":
    app = ConsultaElectoralApp()
    app.mainloop()
